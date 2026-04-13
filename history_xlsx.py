__author__ = "Edgar Bonilla Rivas"
__copyright__ = "Copyright (C) 2025 Author Name"
__license__ = "AUTOMATYCO"
__version__ = "v2.0.0"

import os
import os.path
import pandas as pd
from datetime import date
from datetime import datetime
# import history_query

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

#######################################################################################################################################################

part_name = "P1895152-00-G:SHG2242791000290"

# print(type_station)
def history_file_xlsx(part_name):

    today = date.today()
    today_year = format(today.year)
    today_month = today.strftime("%m")
    today_day = today.strftime("%d")
    date_time = datetime.now()
    date_time = date_time.strftime('%Y-%m-%d %H:%M:%S')

    file_data_bk = "C:/AMC/History/"+today_year+"/"+today_month+"/"+today_day+"/"+today_year+"-"+today_month+"-"+today_day+"_History"+'.xlsx'
    file_data = "C:/Users/Tesla/Documents/Traceability/"+today_year+"/"+today_month+"/"+today_day+"/"+today_year+"-"+today_month+"-"+today_day+"_History"+'.xlsx'

    directory_bk =  "C:/AMC/History/"+today_year+"/"+today_month+"/"+today_day+"/"
    directory =  "C:/Users/Tesla/Documents/Traceability/"+today_year+"/"+today_month+"/"+today_day+"/"

    verify_file_bk = os.path.isfile(file_data_bk)
    verify_file = os.path.isfile(file_data)

    verify_directory_bk = os.path.exists(directory_bk)
    verify_directory = os.path.exists(directory)

    columns_df = ["Part", "Model", "Value","Lower limit","Upper limit","Data type","Unit","Result","Compoperator","Testing time","Metadata","Description", "dwell_time",
                   "Final result - Task result", "Final result - Task time", "Final result - Task duration", "Final result - Metadata"]


    amc = Image("amc.png")
    tesla = Image("tesla.png")

    if verify_file_bk == True:
        # print(part_name)
        try:
            
            # Existing Excel file
            existing_file = file_data_bk

            # New data to append
            new_data = [part_name]
            df_new = pd.DataFrame(new_data)

            # Load existing workbook
            wb = load_workbook(existing_file)

            # Select the active sheet
            ws = wb.active

            # Append new data
            for row in new_data:
                ws.append(row)

            # Save the workbook
            wb.save(existing_file)

        except Exception as e:
            print(str(e))
    else:
        # print(part_name)
        wb = openpyxl.Workbook()
        try:
            if verify_directory_bk != True:
                os.makedirs(directory_bk)

            # Acceder al nombre de la hoja activa y modificarlo
            hoja = wb.active
            hoja.title = "History_"+today_year+"-"+today_month+"-"+today_day

            amc.width = 300
            amc.height = 70

            tesla.width = 150
            tesla.height = 150

            hoja.add_image(amc,'Q2')
            hoja.add_image(tesla,'A1')

            hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['E2'].alignment = Alignment(horizontal='center', vertical='center')

            hoja.column_dimensions['A'].width = 31
            hoja.column_dimensions['A'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['B'].width = 15
            hoja.column_dimensions['B'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['C'].width = 15
            hoja.column_dimensions['C'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['D'].width = 15
            hoja.column_dimensions['D'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['E'].width = 15
            hoja.column_dimensions['E'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['F'].width = 15
            hoja.column_dimensions['F'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['G'].width = 15
            hoja.column_dimensions['G'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['H'].width = 15
            hoja.column_dimensions['H'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['I'].width = 25
            hoja.column_dimensions['I'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['J'].width = 25
            hoja.column_dimensions['J'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['K'].width = 50
            hoja.column_dimensions['K'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['L'].width = 50
            hoja.column_dimensions['L'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['M'].width = 30
            hoja.column_dimensions['M'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['N'].width = 30
            hoja.column_dimensions['N'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['O'].width = 25
            hoja.column_dimensions['O'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['P'].width = 30
            hoja.column_dimensions['P'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['Q'].width = 50
            hoja.column_dimensions['Q'].alignment = Alignment(horizontal='center', vertical='center')

            hoja['G2'] = 'History'
            hoja['A6'] = columns_df[0]
            hoja['B6'] = columns_df[1]
            hoja['C6'] = columns_df[2]
            hoja['D6'] = columns_df[3]
            hoja['E6'] = columns_df[4]
            hoja['F6'] = columns_df[5]
            hoja['G6'] = columns_df[6]
            hoja['H6'] = columns_df[7]
            hoja['I6'] = columns_df[8]
            hoja['J6'] = columns_df[9]
            hoja['K6'] = columns_df[10]
            hoja['L6'] = columns_df[11]
            hoja['M6'] = columns_df[12]
            hoja['N6'] = columns_df[13]
            hoja['O6'] = columns_df[14]
            hoja['P6'] = columns_df[15]
            hoja['Q6'] = columns_df[16]

            hoja['G2'].font = Font(name='Calibri', bold=True, italic=True, size=36, color="000000")
            hoja['G2'].alignment = Alignment(horizontal='center', vertical='center')

            hoja['A6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['B6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['C6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['D6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['E6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['F6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['G6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['H6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['I6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['J6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['K6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['L6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['M6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['N6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['O6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['P6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['Q6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")

            hoja['A6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['B6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['C6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['D6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['E6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['F6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['G6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['H6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['I6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['J6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['K6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['L6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['M6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['N6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['O6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['P6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['Q6'].alignment = Alignment(horizontal='center', vertical='center')

            hoja.append(part_name)

            wb.save(file_data_bk)

        except Exception as e:
            print (str(e))
########################################################### File traceability ########################################################################
    if verify_file == True:
        try:
            # Existing Excel file
            existing_file = file_data

            # New data to append
            new_data = [part_name]
            df_new = pd.DataFrame(new_data)

            # Load existing workbook
            wb = load_workbook(existing_file)

            # Select the active sheet
            ws = wb.active

            # Append new data
            for row in new_data:
                ws.append(row)

            # Save the workbook
            wb.save(existing_file)

        except Exception as e:
            print(str(e))
    else:
        
        wb = openpyxl.Workbook()
        try:
            if verify_directory != True:
                os.makedirs(directory)

            # Acceder al nombre de la hoja activa y modificarlo
            hoja = wb.active
            hoja.title = "History_"+today_year+"-"+today_month+"-"+today_day

            amc.width = 300
            amc.height = 70

            tesla.width = 150
            tesla.height = 150

            hoja.add_image(amc,'Q2')
            hoja.add_image(tesla,'A1')

            hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['E2'].alignment = Alignment(horizontal='center', vertical='center')

            hoja.column_dimensions['A'].width = 31
            hoja.column_dimensions['A'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['B'].width = 15
            hoja.column_dimensions['B'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['C'].width = 15
            hoja.column_dimensions['C'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['D'].width = 15
            hoja.column_dimensions['D'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['E'].width = 15
            hoja.column_dimensions['E'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['F'].width = 15
            hoja.column_dimensions['F'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['G'].width = 15
            hoja.column_dimensions['G'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['H'].width = 15
            hoja.column_dimensions['H'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['I'].width = 25
            hoja.column_dimensions['I'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['J'].width = 25
            hoja.column_dimensions['J'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['K'].width = 50
            hoja.column_dimensions['K'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['L'].width = 50
            hoja.column_dimensions['L'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['M'].width = 30
            hoja.column_dimensions['M'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['N'].width = 30
            hoja.column_dimensions['N'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['O'].width = 25
            hoja.column_dimensions['O'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['P'].width = 30
            hoja.column_dimensions['P'].alignment = Alignment(horizontal='center', vertical='center')
            hoja.column_dimensions['Q'].width = 50
            hoja.column_dimensions['Q'].alignment = Alignment(horizontal='center', vertical='center')

            hoja['G2'] = 'History'
            hoja['A6'] = columns_df[0]
            hoja['B6'] = columns_df[1]
            hoja['C6'] = columns_df[2]
            hoja['D6'] = columns_df[3]
            hoja['E6'] = columns_df[4]
            hoja['F6'] = columns_df[5]
            hoja['G6'] = columns_df[6]
            hoja['H6'] = columns_df[7]
            hoja['I6'] = columns_df[8]
            hoja['J6'] = columns_df[9]
            hoja['K6'] = columns_df[10]
            hoja['L6'] = columns_df[11]
            hoja['M6'] = columns_df[12]
            hoja['N6'] = columns_df[13]
            hoja['O6'] = columns_df[14]
            hoja['P6'] = columns_df[15]
            hoja['Q6'] = columns_df[16]

            hoja['G2'].font = Font(name='Calibri', bold=True, italic=True, size=36, color="000000")
            hoja['G2'].alignment = Alignment(horizontal='center', vertical='center')

            hoja['A6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['B6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['C6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['D6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['E6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['F6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['G6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['H6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['I6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['J6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['K6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['L6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['M6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['N6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['O6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['P6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")
            hoja['Q6'].font = Font(name='Calibri', bold=True, italic=True, size=14, color="FF6347")

            hoja['A6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['B6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['C6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['D6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['E6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['F6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['G6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['H6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['I6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['J6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['K6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['L6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['M6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['N6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['O6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['P6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['Q6'].alignment = Alignment(horizontal='center', vertical='center')

            hoja.append(part_name)

            wb.save(file_data)

        except Exception as e:
            print (str(e)) 

# history_file_xlsx(part_name)